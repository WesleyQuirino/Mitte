import flet
from flet import (
    ElevatedButton,
    FilePicker,
    FilePickerResultEvent,
    Page,
    Row,
    Text,
    icons,
)
import tabula
import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import sys

# Função para processar o PDF e gerar JSON e Excel
def process_pdf(pdf_path):
    try:

        # Extrair o nome do arquivo PDF sem a extensão
        file_name = os.path.splitext(os.path.basename(pdf_path))[0]

        # Verificar e criar pastas json e excel, se necessário
        # os.makedirs('./json', exist_ok=True)
        # os.makedirs('./excel', exist_ok=True)

        # Extrair todas as tabelas do PDF
        tables = tabula.read_pdf(
            pdf_path,
            pages="all",
            multiple_tables=True,
            relative_area=True,
            relative_columns=True,
            area=[12, 0, 90, 100],
            columns=[10, 30.5, 37, 45, 53, 60, 70, 80, 100]
        )

        if not tables or len(tables) == 0:
            raise ValueError("Nenhuma tabela foi extraída do PDF.")

        custom_header = ["ID", "Nome", "Bloco", "Apto", "Dispositivo", "Saída", "Recurso", "Status do Recurso", "Data de registro"]
        processed_tables = []

        for table in tables:
            table.columns = custom_header
            table = table.drop(table.index[0:2])
            processed_tables.append(table)

        final_table = pd.concat(processed_tables, ignore_index=True)
        final_table = final_table.where(pd.notnull(final_table), None)
        json_data = final_table.to_dict(orient="records")

        processed_data = []
        previous_item = {}

        for i in range(len(json_data)):
            if not json_data[i].get("ID"):
                for json_data_key, json_data_value in json_data[i].items():
                    if json_data_value:
                        if previous_item.get(json_data_key):
                            if json_data_value not in previous_item[json_data_key]:
                                previous_item[json_data_key] += " " + json_data_value
                        else:
                            previous_item[json_data_key] = json_data_value
                processed_data[-1] = previous_item
            else:
                previous_item = json_data[i]
                processed_data.append(previous_item)

        for i in range(len(processed_data)):
            data_registro = processed_data[i]["Data de registro"]
            data, hora = data_registro.split(" ")
            processed_data[i]["ID"] = int(processed_data[i]["ID"])
            if processed_data[i]["Apto"] != None:
                processed_data[i]["Apto"] = int(processed_data[i]["Apto"])
            processed_data[i]["Data"] = data
            processed_data[i]["Hora"] = hora
            del processed_data[i]["Data de registro"]

        json_string = json.dumps(processed_data, ensure_ascii=False, indent=4)

        # Caminhos dos arquivos JSON e Excel
        # json_output_path = f"./json/{file_name}.json"
        excel_output_path = f"./{file_name}.xlsx"

        # Salvar JSON
        # with open(json_output_path, "w", encoding="utf-8") as json_file:
        #     json_file.write(json_string)

        for i in range(len(processed_data)):
            processed_data[i]["Data"] = datetime.strptime(processed_data[i]["Data"], "%d/%m/%Y")
            processed_data[i]["Hora"] = datetime.strptime(processed_data[i]["Hora"], "%H:%M:%S")

        processed_df = pd.DataFrame(processed_data)
        processed_df.to_excel(excel_output_path, index=False)

        wb = load_workbook(excel_output_path)
        ws = wb.active

        column_widths = {
            "A": 10,
            "B": 25,
            "C": 10,
            "D": 10,
            "E": 20,
            "F": 25,
            "G": 25,
            "H": 25,
            "I": 20,
            "J": 15,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill

        for row in range(2, ws.max_row + 1):
            if row % 2 != 0:
                for cell in ws[row]:
                    cell.fill = header_fill

        ws.auto_filter.ref = ws.dimensions
        # Definir formato de data e número
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Começar da segunda linha
            for cell in row:
                if cell.column_letter == 'I':  # Supondo que a coluna I contém as datas
                    cell.number_format = 'DD/MM/YYYY'  # Formato de data
                elif cell.column_letter == 'J':  # Supondo que a coluna J contém horas
                    cell.number_format = 'HH:MM:SS'  # Formato de hora
                elif cell.column_letter in ['A', 'D']:  # Supondo que essas colunas são numéricas
                    cell.number_format = '0'  # Formato numérico
        ws.sheet_view.showGridLines = False
        wb.save(excel_output_path)
        # return json_output_path
        return excel_output_path
    except Exception as e:
        sys.stderr.write(f"Erro ao processar o PDF: {str(e)}\n")
        return None

# Interface com Flet
def main(page: Page):
    def pick_files_result(e: FilePickerResultEvent):
        if e.files:
            pdf_file = e.files[0].path
            selected_files.value = f"Arquivo selecionado: {e.files[0].name}"
            selected_files.update()

            # Processa o PDF
            excel_file = process_pdf(pdf_file)
            if excel_file:
                # download_json_link.value = f"Baixar JSON: {os.path.basename(json_file)}"
                download_excel_link.value = f"Baixar Excel: {os.path.basename(excel_file)}"
                # download_json_link.update()
                download_excel_link.update()
            else:
                selected_files.value = "Erro ao processar o PDF"
                selected_files.update()

    pick_files_dialog = FilePicker(on_result=pick_files_result)
    selected_files = Text()

    # download_json_link = Text(value="")
    download_excel_link = Text(value="")

    page.overlay.append(pick_files_dialog)

    page.add(
        Row(
            [
                ElevatedButton(
                    "Selecionar PDF",
                    icon=icons.UPLOAD_FILE,
                    on_click=lambda _: pick_files_dialog.pick_files(allow_multiple=False),
                ),
                selected_files,
            ]
        ),
        # Row([download_json_link]),
        Row([download_excel_link]),
    )

flet.app(target=main)