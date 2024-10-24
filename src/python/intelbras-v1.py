import tabula
import pandas as pd
import json
import sys
import os  # Para trabalhar com o nome do arquivo
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

def process_pdf(pdf_path):
    try:
        # Extrair o nome do arquivo PDF sem a extensão
        file_name = os.path.splitext(os.path.basename(pdf_path))[0]

        # Extraindo todas as tabelas do PDF
        tables = tabula.read_pdf(pdf_path, 
            pages="all", 
            multiple_tables=True, 
            relative_area=True, 
            relative_columns=True, 
            area=[12,0,90,100], 
            columns=[10, 30.5, 37, 45, 53, 60, 70, 80, 100]
        )

        if not tables or len(tables) == 0:
            raise ValueError("Nenhuma tabela foi extraída do PDF.")

        # Definir o cabeçalho personalizado
        custom_header = ["ID", "Nome", "Bloco", "Apto", "Dispositivo", "Saída", "Recurso", "Status do Recurso", "Data de registro"]

        # Lista para armazenar todas as tabelas processadas
        processed_tables = []

        # Iterar sobre todas as tabelas extraídas
        for table in tables:
            # Definir o cabeçalho personalizado para cada tabela
            table.columns = custom_header

            # Remover as linhas com índices de 0 a 1
            table = table.drop(table.index[0:2])

            # Adicionar a tabela processada à lista
            processed_tables.append(table)

        # Concatenar todas as tabelas em um único DataFrame
        final_table = pd.concat(processed_tables, ignore_index=True)

        # Substituir NaN por None (que será convertido para null no JSON)
        final_table = final_table.where(pd.notnull(final_table), None)

        # Converte o DataFrame para uma lista de dicionários (JSON)
        json_data = final_table.to_dict(orient='records')

        # Processar os dados
        processed_data = []
        previous_item = {}

        for i in range(len(json_data)):
            if not json_data[i].get("ID"):
                for json_data_key, json_data_value in json_data[i].items():
                    if json_data_value:
                        if previous_item.get(json_data_key):
                            if json_data_value not in previous_item[json_data_key]:
                                previous_item[json_data_key] += " " + json_data_value
                        else:
                            previous_item[json_data_key] = json_data_value
                processed_data[-1] = previous_item
            else:
                previous_item = json_data[i]
                processed_data.append(previous_item)

        for i in range(len(processed_data)):
            data_registro = processed_data[i]["Data de registro"]

            # Divide a data e a hora
            data, hora = data_registro.split(' ')

            # Atualiza o objeto com as novas chaves "Data" e "Hora"
            processed_data[i]["Data"] = data
            processed_data[i]["Hora"] = hora

            # Remove a chave original "Data de registro"
            del processed_data[i]["Data de registro"]

        # Converte a lista final para JSON formatado
        json_string = json.dumps(processed_data, ensure_ascii=False, indent=4)

        # Criar o nome do arquivo JSON e Excel usando o nome do PDF
        json_output_path = f'./json/{file_name}.json'
        excel_output_path = f'./excel/{file_name}.xlsx'

        # Exportar os dados processados para um arquivo JSON
        with open(json_output_path, 'w', encoding='utf-8') as json_file:
            json_file.write(json_string)

        for i in range(len(processed_data)):
            # Atualiza o objeto com as novas chaves "Data" e "Hora"
            processed_data[i]["Data"] = datetime.strptime(processed_data[i]["Data"], "%d/%m/%Y")
            processed_data[i]["Hora"] = datetime.strptime(processed_data[i]["Hora"], "%H:%M:%S")
        
        # Converter o processed_data para um DataFrame e exportar para Excel
        processed_df = pd.DataFrame(processed_data)
        processed_df.to_excel(excel_output_path, index=False)

        # Carregar o arquivo Excel para aplicar estilos e definir larguras de colunas
        wb = load_workbook(excel_output_path)
        ws = wb.active

        # Definir largura fixa para as colunas
        column_widths = {
            "A": 10,
            "B": 25,
            "C": 10,
            "D": 10,
            "E": 20,
            "F": 25,
            "G": 25,
            "H": 25,
            "I": 20,
            "J": 15,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Aplicar estilo de fundo cinza claro para o cabeçalho
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        for cell in ws[1]:  # Assumindo que o cabeçalho está na primeira linha
            cell.fill = header_fill

        # Aplicar estilo de fundo cinza claro para linhas ímpares
        for row in range(2, ws.max_row + 1):  # Começa na linha 2 para pular o cabeçalho
            if row % 2 != 0:  # Linhas ímpares
                for cell in ws[row]:
                    cell.fill = header_fill

        # Aplicar filtro no cabeçalho
        ws.auto_filter.ref = ws.dimensions

        # Definir formato de data e número
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Começar da segunda linha
            for cell in row:
                if cell.column_letter == 'I':  # Supondo que a coluna I contém as datas
                    cell.number_format = 'DD/MM/YYYY'  # Formato de data
                elif cell.column_letter == 'J':  # Supondo que a coluna J contém horas
                    cell.number_format = 'HH:MM:SS'  # Formato de hora
                elif cell.column_letter in ['A', 'E', 'F', 'G']:  # Supondo que essas colunas são numéricas
                    cell.number_format = '0'  # Formato numérico

        # Remover as linhas de grade
        ws.sheet_view.showGridLines = False

        # Salvar as alterações
        wb.save(excel_output_path)

        # Retornar o JSON processado como string
        return json_string
    except Exception as e:
        # Redirecionar a mensagem de erro para o stderr
        sys.stderr.write(f"Erro ao processar o PDF: {str(e)}\n")
        return None

# Receber o caminho do PDF como argumento
if __name__ == "__main__":
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
        result = process_pdf(pdf_path)

        if result:
            # Se o JSON foi gerado com sucesso, exibir no stdout
            print(result)
        else:
            # Se houve erro, retornar uma mensagem clara
            sys.stderr.write("Erro ao processar o JSON.\n")
    else:
        sys.stderr.write("Por favor, forneça o caminho para o arquivo PDF como argumento.\n")